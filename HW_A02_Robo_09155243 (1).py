# -*- coding: utf-8 -*-
"""
Created on Fri Dec  2 10:00:10 2022

@author: ant33
"""
#1.年化報酬率、報酬率標準差、報酬率偏態係數、報酬率峰態係數)，以及每年100檔台股的相關係數 ρij
import yfinance as yf
import pandas as pd
import numpy as np
from scipy import stats
import yfinance as yf
import pandas as pd
import numpy as np
import datetime as dt
import statsmodels.api as sm
rf=0.0
df=pd.read_html('https://www.taifex.com.tw/cht/9/futuresQADetail')
df=df[0]
df=df[['證券名稱','證券名稱.1']]
df=df.iloc[0:100]
start=dt.datetime(2006,1,2)
end=dt.datetime(2022,12,31)
ticker=[str(i)+'.TW' for i in df['證券名稱']]
df_yf=yf.download(ticker,start=start,end=end)
n_close=len(df_yf['Adj Close'].columns)
df_yf1=df_yf['Adj Close'][ticker]
df_yf1.fillna(method='ffill',inplace=True)

for i in range(start.year,end.year+1):
    i=str(i)
    vars()["a_"+i]=df_yf1[i].dropna(thresh=10,axis=1)
    vars()["return_"+i]=np.log(vars()["a_"+i]/vars()["a_"+i].shift(1))
    vars()["return_"+i].drop(labels=vars()["return_"+i].index[0],axis=0,inplace=True)
    vars()["return_"+i].fillna(method='ffill',inplace=True)
    vars()["mean_"+i]=np.mean(vars()["return_"+i])*250
    vars()["std_"+i]=np.std(vars()["return_"+i])*(np.sqrt(250))
    vars()["skew_"+i]=vars()["return_"+i].skew()
    vars()["kurt_"+i]=vars()["return_"+i].kurt()
    vars()["corr_"+i]=vars()["return_"+i].corr()


#1.capm(α、β)
import yfinance as yf
import pandas as pd
import numpy as np
import datetime as dt
import statsmodels.api as sm
rf=0.0
df=pd.read_html('https://www.taifex.com.tw/cht/9/futuresQADetail')
df=df[0]
df=df[['證券名稱','證券名稱.1']]
df=df.iloc[0:100]
start=dt.datetime(2006,1,2)
end=dt.datetime(2022,12,31)
ticker1=[str(i)+'.TW' for i in df['證券名稱']]
ticker1.append('^TWII')
df_yf2=yf.download(ticker1,start=start,end=end)
n_close1=len(df_yf2['Adj Close'].columns)
df_yf2=df_yf2['Adj Close'][ticker1]
df_yf2.fillna(method='ffill',inplace=True)

def CAPM(df_return):
    capm_out = {}
    for company in df_return.columns[:n_close-1]:
        if df_return[company].isna().sum()<len(df_return.index):
           df_return[company]=df_return[company]-rf 
           df_return['^TWII']=df_return['^TWII']-rf
           data_df=pd.concat([df_return[company],df_return['^TWII']],axis=1)
           data_df.columns=[company,'index']
           X=data_df[['index']].assign(Intercept=1)
           Y=data_df[company]
           X=sm.add_constant(X)
           model=sm.OLS(Y,X)
           results=model.fit()
           Alpha=results.params[1]
           Beta=results.params[0]
           capm_out[company]=[Alpha, Beta]
        else:
           Alpha=np.nan
           Beta=np.nan
           capm_out[company]=[Alpha,Beta]
          
    df_ab=pd.DataFrame.from_dict(capm_out, orient='index',columns=['Alpha','Beta'])
    return df_ab

for i in range(start.year,end.year+1):
    i=str(i)
    vars()["a_"+i]=df_yf2[i].dropna(thresh=50,axis=1)
    vars()["return_"+i]=np.log(vars()["a_"+i]/vars()["a_"+i].shift(1))
    vars()["return_"+i].drop(labels=vars()["return_"+i].index[0],axis=0,inplace=True)
    vars()["return_"+i].fillna(method='ffill',inplace=True)
    vars()["CAPM_"+i]=CAPM(vars()["return_"+i])



writer = pd.ExcelWriter("HW_A02_Robo_09155243.xlsx",engine = 'xlsxwriter')

for i in range(start.year,end.year+1):
    i = str(i)
    vars()["CAPM_"+i].to_excel(writer,sheet_name = "CAPM_"+i)
writer.save()

'''
from openpyxl import load_workbook #為了第一格
wb = load_workbook('HW_A02_Robo_09155243.xlsx')
for i in wb:
    ws = wb[i]
    ws['A1'] = '股票代碼'
wb.save('HW_A02_Robo_09155243.xlsx')
'''

#2找到前三名   max alpha 最小標準差
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt

data = pd.read_excel('HW_A02_Robo_09155243.xlsx', sheet_name = None)
for i in data.keys():
    print(i,data[i])
    j=data[i]
    vars()["Alpha_sort_"+i]=j.sort_values('Alpha', ascending=False).head(3)
    vars()["Alpha_sort_"+i].columns=['股票代碼','Alpha','Beta']
'''
    for j in vars()["Alpha_sort_"+i]['股票代碼']:
        for k in range(start.year+1,end.year+1):
            k=str(k)
            vars()["Alpha_stock_"+k]= yf.download(j,start=k+'-01-01',end=k+'-12-31')
            vars()["Alpha_stock_"+k]=vars()["Alpha_stock_"+k]['Adj Close']
            vars()["Alpha_stock_return_"+i]=np.log(vars()["Alpha_stock_"+k]/vars()["Alpha_stock_"+k].shift(1))
            vars()["Alpha_stock_mean"+k]=np.mean(vars()["Alpha_stock_return_"+i])*250
            vars()["Alpha_stock_std"+k]=np.std(vars()["Alpha_stock_return_"+i])*(np.sqrt(250))
            
'''      
'''
for k in range(start.year+1,end.year+1):
    k=str(k)
    vars()['b_'+k]=[]
    vars()['b_'+k].append(vars()["Alpha_sort_CAPM_"+k]['股票代碼'])
    for j in vars()['b_'+k]:
        j.replace(',','')
        vars()["Alpha_stock_"+k]= yf.download(j,start=k+'-01-01',end=k+'-12-31')
'''
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
from pandas_datareader import data as pdr

ticker2007=['3702.TW','2347.TW','2360.TW']
ticker2008=['2880.TW','2885.TW','2881.TW']
ticker2009=['3533.TW','2360.TW','3702.TW']
ticker2010=['2049.TW','2618.TW','2609.TW']
ticker2011=['2049.TW','2207.TW','2474.TW']
ticker2012=['1476.TW','5871.TW','9945.TW']      
ticker2013=['1476.TW','2356.TW','9910.TW']     
ticker2014=['3533.TW','2327.TW','4938.TW']
ticker2015=['8464.TW','2345.TW','9910.TW']
ticker2016=['2633.TW','2371.TW','1590.TW']
ticker2017=['2327.TW','3443.TW','3661.TW']
ticker2018=['6669.TW','2027.TW','3037.TW']
ticker2019=['3661.TW','2207.TW','6669.TW']
ticker2020=['2609.TW','8046.TW','2603.TW']
ticker2021=['8454.TW','2615.TW','2609.TW']
ticker2022=['1605.TW','3443.TW','3533.TW']

 
for k in range(start.year+1,end.year+1):
    k=str(k)
    vars()["Alpha_stock_"+k]= yf.download(vars()["ticker"+k],start=k+'-01-01',end=k+'-12-31')
    vars()["Alpha_stock_"+k]=vars()["Alpha_stock_"+k]['Adj Close']
    vars()["Alpha_stock_"+k].fillna(method='ffill',inplace=True)
    vars()["Alpha_stock_return_"+k]=np.log(vars()["Alpha_stock_"+k]/vars()["Alpha_stock_"+k].shift(1))
    vars()["Alpha_stock_return_"+k].drop(labels=vars()["Alpha_stock_return_"+k].index[0],axis=0,inplace=True)
    vars()["Alpha_stock_mean"+k]=np.mean(vars()["Alpha_stock_return_"+k])*250
    vars()["Alpha_stock_std"+k]=np.std(vars()["Alpha_stock_return_"+k])*(np.sqrt(250))
    vars()["Alpha_stock_corr"+k]=vars()["Alpha_stock_return_"+k].corr()
    vars()["Alpha_stock_cov"+k]=vars()["Alpha_stock_return_"+k].cov()
    vars()["Alpha_stock_num_assets"+k] = len(vars()["Alpha_stock_"+k].columns)
    num_portfolios = 10000
    for portfolio in range(num_portfolios):
        vars()["Alpha_stock_weights"+k] = np.random.random(vars()["Alpha_stock_num_assets"+k])
        vars()["Alpha_stock_weights"+k] = vars()["Alpha_stock_weights"+k]/np.sum(vars()["Alpha_stock_weights"+k])
    vars()["Alpha_stock_year_return"+k]=vars()["Alpha_stock_mean"+k][0]*vars()["Alpha_stock_weights"+k][0]+vars()["Alpha_stock_mean"+k][1]*vars()["Alpha_stock_weights"+k][1]+vars()["Alpha_stock_mean"+k][2]*vars()["Alpha_stock_weights"+k][2]
    vars()["Alpha_stock_year_std"+k]=vars()["Alpha_stock_std"+k][0]*vars()["Alpha_stock_weights"+k][0]+vars()["Alpha_stock_std"+k][1]*vars()["Alpha_stock_weights"+k][1]+vars()["Alpha_stock_std"+k][2]*vars()["Alpha_stock_weights"+k][2]
    vars()["Alpha_stock_year_sharpe"+k]=(vars()["Alpha_stock_year_return"+k]-rf)/vars()["Alpha_stock_year_std"+k]

for k in range(start.year+1,end.year+1):
    k=str(k)    
    print(k+'年的報酬,報酬標準差,夏普比率為:',vars()["Alpha_stock_year_return"+k], vars()["Alpha_stock_year_std"+k] , vars()["Alpha_stock_year_sharpe"+k])
        
            
            